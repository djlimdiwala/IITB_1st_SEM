import csv
import sys
import xlsxwriter
import re
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


exampleFile = open(sys.argv[1] + ".csv")
exampleReader = csv.reader(exampleFile)
data = list(exampleReader)



wb = Workbook()
dest_filename = sys.argv[1] + "_" + sys.argv[2] + ".xlsx"
p = string.ascii_uppercase
m = len(data)
n = len(data[0])
counter = 0
flag = 0
col = 0
header = data[0]
for i in range(m):
	for j in range(n):
		if counter == 0:
			obj = re.search(r'(.*?)(work\sno)(.*?)',data[i][j].lower())
			if obj:
				flag = 1
				col = j
				break
				counter = counter + 1

if flag == 0:
	print ("No header containing work no attribute")
	sys.exit()
counter = 1
for i in range(m):
	if data[i][col] == sys.argv[2]:
		sh = sys.argv[2] + "_" + str(counter)
		ws2 = wb.create_sheet(title = sh)
		font = Font(size=13, bold = True)
		font1 = Font(size=10, bold = False)
		for k in range(n):
			
			ws2[str(p[k] + str(1))] = header[k]

			if k == col:
				ws2[str(p[k] + str(2))] = str(data[i][k]) + "_" + str(counter)

			else:
				ws2[str(p[k] + str(2))] = data[i][k]
		counter = counter + 1

		for cell in ws2["1:1"]:
			cell.font = font
		for cell in ws2["2:2"]:
			cell.font = font1


if counter == 1:
	print ("\n" + "No record with given work no found ...")
	sys.exit()
std=wb.get_sheet_by_name('Sheet')
wb.remove_sheet(std)


wb.save(filename = dest_filename)
