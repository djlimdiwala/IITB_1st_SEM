import pandas as pd
import numpy as np
import xlsxwriter
import re
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlsxwriter
import matplotlib.pyplot as plt
from pylab import *

# create workbook for writing results in excel file
wb = Workbook()
dest_filename = "173050061_solution.xls"




# Answer to question 1
# Top five import and export destinations, by total imports and total exports
def q_1():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()

	# apply groupby opertion on countries, sort them and take top five (by head function)
	topI = dfi.groupby('Country')['Value-INR-2011-12'].sum().sort_values(ascending=False)
	topE = dfe.groupby('Country')['Value-INR-2011-12'].sum().sort_values(ascending=False)

	p1 = topE.index.values
	p2 = list(topE)
	pp = topI.index.values.size
	ppp = np.arange(pp)
	ind = np.arange(5)
	plt.pie(p2,labels=p1)
	plt.show()
	p1 = topE.index.values
	p2 = list(topE)
	pp = topE.index.values.size
	ppp = np.arange(pp)
	ind = np.arange(5)
	plt.pie(p2,labels=p1)
	plt.show()
	topI = topI.head()
	topE = topE.head()
	p1 = topI.index.values
	p2 = list(topI)
	pp = topI.index.values.size
	ppp = np.arange(pp)
	ind = np.arange(5)
	topI.columns = ['imports']
	topE.columns = ['exports']
	# print (topI)
	# print (topE)
	
	# print (p1)
	# print (p2)
	width = 0.35
	plt.bar(ppp,p2,align='center', alpha=0.35)
	plt.xticks(ppp, p1)
	plt.title("Top five coutries wrt total imports")
	plt.show()



	p1 = topE.index.values
	p2 = list(topE)
	pp = topE.index.values.size
	ppp = np.arange(pp)
	ind = np.arange(5)
	width = 0.35
	plt.bar(ppp,p2,align='center', alpha=0.35)
	plt.xticks(ppp, p1)
	plt.title("Top five coutries wrt total exports")
	plt.show()

	# write results to sheets in excel file
	pp = topI.index.values
	name = "Q1_Imports"
	font = Font(size=13, bold = True)
	ws1 = wb.create_sheet(title = name)
	ws1['A1'] = "Country"
	ws1['A2'] = pp[0]
	ws1['A3'] = pp[1]
	ws1['A4'] = pp[2]
	ws1['A5'] = pp[3]
	ws1['A6'] = pp[4]
	for cell in ws1["1:1"]:
		cell.font = font
	pp = topE.index.values
	ws1.column_dimensions['A'].width = 15
	name = "Q1_Exports"
	font = Font(size=13, bold = True)
	ws1 = wb.create_sheet(title = name)
	ws1['A1'] = "Country"
	ws1['A2'] = pp[0]
	ws1['A3'] = pp[1]
	ws1['A4'] = pp[2]
	ws1['A5'] = pp[3]
	ws1['A6'] = pp[4]
	for cell in ws1["1:1"]:
		cell.font = font

	ws1.column_dimensions['A'].width = 15




# Answer to question 2
# Top five import and export commodities
def q_2():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()

	# apply groupby opertion on commodities, sort them and take top five (by head function)
	coI = dfi.groupby('Commodity')['Value-INR-2011-12'].sum().sort_values(ascending=False).head()
	coE = dfe.groupby('Commodity')['Value-INR-2011-12'].sum().sort_values(ascending=False).head()

	# print (coI)
	# print (coE)

	# write results to sheets in excel file
	pp = coI.index.values
	name = "Q2_Imports"
	font = Font(size=13, bold = True)
	ws1 = wb.create_sheet(title = name)
	ws1['A1'] = "Commodity"
	ws1['A2'] = pp[0]
	ws1['A3'] = pp[1]
	ws1['A4'] = pp[2]
	ws1['A5'] = pp[3]
	ws1['A6'] = pp[4]
	for cell in ws1["1:1"]:
		cell.font = font
	ws1.column_dimensions['A'].width = 35
	pp = coE.index.values
	name = "Q2_Exports"
	font = Font(size=13, bold = True)
	ws1 = wb.create_sheet(title = name)
	ws1['A1'] = "Commodity"
	ws1['A2'] = pp[0]
	ws1['A3'] = pp[1]
	ws1['A4'] = pp[2]
	ws1['A5'] = pp[3]
	ws1['A6'] = pp[4]
	for cell in ws1["1:1"]:
		cell.font = font
	ws1.column_dimensions['A'].width = 35




# Answer to question 3
# Total imports, total exports, export/import ratio, export-import (trade deficit) for each country
def q_3():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	p = string.ascii_uppercase
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()


	# apply groupby opertion on countries and aggregate on sum
	topI = dfi.groupby('Country')['Value-INR-2011-12'].sum()
	topE = dfe.groupby('Country')['Value-INR-2011-12'].sum()

	# concatenate import and export columns horizontally
	cat = pd.concat([topI, topE],axis=1)
	cat.columns = ['Total_Imports','Total_Exports']
	# cacculate Export/import ratio and trade deficit
	cat['Export/import ratio'] = cat['Total_Exports'] / cat['Total_Imports']
	cat['Export-import (trade deficit)'] = cat['Total_Exports'] - cat['Total_Imports']
	# print (cat)
	# print (merge)
	# result = cat.to_latex()
	# print (result)
	# f1 = open("table1.tex",'w')
	# f1.write(result)
	# f1.close()
	# write results to sheets in excel file
	pp = cat.index.values
	n = cat.index.values.size
	p1 = cat['Total_Imports']
	p2 = cat['Total_Exports']
	p3 = cat['Export/import ratio']
	p4 = cat['Export-import (trade deficit)']

	ws1 = wb.create_sheet(title = "Q3")
	ws1['A1'] = "Country"
	ws1['B1'] = "Total_Imports"
	ws1['C1'] = "Total_Exports"
	ws1['D1'] = "Export/import ratio"
	ws1['E1'] = "Export-import (trade deficit)"
	font = Font(size = 13, bold = True)
	for cell in ws1["1:1"]:
		cell.font = font
	for k in range(n):	
		i = 0
		ws1[str(p[i] + str(k+2))] = pp[k]
		i = i + 1
		ws1[str(p[i] + str(k+2))] = p1[k]
		i = i + 1
		ws1[str(p[i] + str(k+2))] = p2[k]
		i = i + 1
		ws1[str(p[i] + str(k+2))] = p3[k]
		i = i + 1
		ws1[str(p[i] + str(k+2))] = p4[k]
		i = i + 1

	ws1.column_dimensions['A'].width = 20
	ws1.column_dimensions['B'].width = 25
	ws1.column_dimensions['C'].width = 25
	ws1.column_dimensions['D'].width = 25
	ws1.column_dimensions['E'].width = 25




# Answer to question 4
# All countries to whom our export is more than Rs 10,000 Cr using query method 
def q_4():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()
	p = string.ascii_uppercase

	# Apply group by operation on Country and sum on aggregate
	topI = dfi.groupby('Country')['Value-INR-2011-12'].sum()
	topE = dfe.groupby('Country')['Value-INR-2011-12'].sum()

	# concatenate import and export columns horizontally
	merge = pd.concat([topE, topI],axis=1)
	merge.columns = ['exports', 'imports']

	# query opration to find countries whose export is more than 10,000 Cr
	merge = merge.query('exports > 1e+11')
	# print (merge)
	result = merge.to_latex()
	# print (result)
	f1 = open("table1.tex",'w')
	f1.write(result)
	f1.close()
	# write results to sheets in excel file
	font = Font(size = 13, bold = True)
	pp = merge.index.values
	n = merge.index.values.size
	ws1 = wb.create_sheet(title = "Q4")
	ws1['A1'] = "Country"
	for k in range(n):
		ws1[str(p[0] + str(k+2))] = pp[k]
	for cell in ws1["1:1"]:
		cell.font = font
	ws1.column_dimensions['A'].width = 20



# Answer to question 5
# Renaming the columns of the answer in question 4 to: 'Country', 'Exports', 'Imports
def q_5():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	p = string.ascii_uppercase
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()

	# Apply group by operation on Country and sum on aggregate
	topI = dfi.groupby('Country')['Value-INR-2011-12'].sum()
	topE = dfe.groupby('Country')['Value-INR-2011-12'].sum()

	# concatenate import and export columns horizontally
	merge = pd.concat([topE, topI],axis=1)
	merge.columns = ['exports', 'imports']

	# query opration to find countries whose export is more than 10,000 Cr
	merge = merge.query('exports > 1e+11')

	# create dataframe temp with index from 0 to n-1 and one column with values as n countries
	temp = pd.DataFrame(merge.index.values, index = np.arange(merge.index.values.size))

	# Concatenate newly created dataframe temp with merge horizontally and set index 0 to n-1 for new concatenated dataframe
	merge = merge.set_index(np.arange(merge.index.values.size))
	merge = pd.concat([temp, merge],axis=1)
	merge.columns = ['Country', 'Exports', 'Imports']
	# print (merge)


	# write results to sheets in excel file
	pp = merge['Country']
	n = merge.index.values.size
	p1 = merge['Exports']
	p2 = merge['Imports']
	ws1= wb.create_sheet(title = "Q5")
	ws1['A1'] = "Country"
	ws1['B1'] = "Import(INR)"
	ws1['C1'] = "Export(INR)"
	font = Font(size = 13, bold = True)
	for cell in ws1["1:1"]:
		cell.font = font
	for k in range(n):
		ws1[str(p[0] + str(k+2))] = pp[k]
		ws1[str(p[1] + str(k+2))] = p1[k]
		ws1[str(p[2] + str(k+2))] = p2[k]
	ws1.column_dimensions['A'].width = 20
	ws1.column_dimensions['B'].width = 25
	ws1.column_dimensions['C'].width = 25


# Answer to question 6
# Creating a new table with column headings "Country", "Transaction", "Value" from table in answer 5 using melt mothod.
def q_6():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()
	p = string.ascii_uppercase

	# Apply group by operation on Country and sum on aggregate
	topI = dfi.groupby('Country')['Value-INR-2011-12'].sum()
	topE = dfe.groupby('Country')['Value-INR-2011-12'].sum()

	# concatenate import and export columns horizontally
	merge = pd.concat([topE, topI],axis=1)
	merge.columns = ['exports', 'imports']

	# query opration to find countries whose export is more than 10,000 Cr
	merge = merge.query('exports > 1e+11')

	# create dataframe temp with index from 0 to n-1 and one column with values as n countries
	temp = pd.DataFrame(merge.index.values, index = np.arange(merge.index.values.size))

	# Concatenate newly created dataframe temp with merge horizontally and set index 0 to n-1 for new concatenated dataframe
	merge = merge.set_index(np.arange(merge.index.values.size))
	merge = pd.concat([temp, merge],axis=1)
	merge.columns = ['Country', 'Exports', 'Imports']

	# Use melt function to reform dataframe intp  the form -- Country, Transaction, Value(INR)
	melted = pd.melt(merge, id_vars = ['Country'], var_name = 'Transaction', value_name = 'Value').sort_values(['Value'],ascending = False).head(n=10)
	# print (melted)

	result = melted.to_latex()
	# print (result)
	f1 = open("table2.tex",'w')
	f1.write(result)
	f1.close()
	# write results to sheets in excel file
	p1 = melted['Country']
	p2 = melted['Transaction']
	p3 = melted['Value']
	index = melted.index.values
	ws1 = wb.create_sheet(title = "Q6")
	ws1['A1'] = "Country"
	ws1['B1'] = "Transaction"
	ws1['C1'] = "Value(INR)"
	font = Font(size = 13, bold = True)
	for cell in ws1["1:1"]:
		cell.font = font
	n = 10
	for k in range(n):
		ws1[str(p[0] + str(k+2))] = p1[index[k]]
		ws1[str(p[1] + str(k+2))] = p2[index[k]]
		ws1[str(p[2] + str(k+2))] = p3[index[k]]
	ws1.column_dimensions['A'].width = 20
	ws1.column_dimensions['B'].width = 10
	ws1.column_dimensions['C'].width = 25



# Answer to question 7
# Commodities that we both export and import.
def q_7():
	# Load import and export files into pandas dataframes
	xls_file = pd.ExcelFile('India_Imports_2011-12_And_2012-13.xls')
	dfi = xls_file.parse()
	p = string.ascii_uppercase
	xls_file = pd.ExcelFile('India_Exports_2011-12_And_2012-13.xls')
	dfe = xls_file.parse()


	# Apply group by operation on Commodity, sum on aggregate and sort them as per Value(INR)
	co12I = dfi.groupby('Commodity')['Value-INR-2011-12'].sum().sort_values(ascending=False)
	co12E = dfe.groupby('Commodity')['Value-INR-2011-12'].sum().sort_values(ascending=False)
	co13I = dfi.groupby('Commodity')['Value-INR-2012-13'].sum().sort_values(ascending=False)
	co13E = dfe.groupby('Commodity')['Value-INR-2012-13'].sum().sort_values(ascending=False)

	# concatenate import and export columns horizontally for each year 2011-12 and 2012-13 and store them in merge12 and merge13 respectively
	merge12 = pd.concat([co12E, co12I],axis=1)
	merge12.columns = ['exports', 'imports']
	temp = pd.DataFrame(merge12.index.values, index = np.arange(merge12.index.values.size))
	merge12 = merge12.set_index(np.arange(merge12.index.values.size))
	merge12 = pd.concat([temp, merge12],axis=1)
	merge12.columns = ['Commodity', 'Exports', 'Imports']

	merge13 = pd.concat([co13E, co13I],axis=1)
	merge13.columns = ['exports', 'imports']
	temp = pd.DataFrame(merge13.index.values, index = np.arange(merge13.index.values.size))
	merge13 = merge13.set_index(np.arange(merge13.index.values.size))
	merge13 = pd.concat([temp, merge13],axis=1)
	merge13.columns = ['Commodity', 'Exports', 'Imports']

	# concatenate merge12 and merge13 vertically
	merge = pd.concat([merge12, merge13],axis=0)
	merge = merge.groupby('Commodity').sum()
	merge = merge.query('Exports > 0')
	merge = merge.query('Imports > 0')
	# print (merge)

	result = merge.to_latex()
	# print (result)
	f1 = open("table3.tex",'w')
	f1.write(result)
	f1.close()
	# write results to sheets in excel file
	p1 = merge.index.values
	n = merge.index.values.size
	ws1 = wb.create_sheet(title = "Q7")
	ws1['A1'] = "Country"
	font = Font(size = 13, bold = True)
	ws1['A1'].font = font
	for k in range(n):
		ws1[str(p[0] + str(k+2))] = p1[k]
	ws1.column_dimensions['A'].width = 35








# Executing all functions
q_1()
q_2()
q_3()
q_4()
q_5()
q_6()
q_7()

data = np.random.randn(2000)
plt.hist(data)
plt.title("Gaussian Histogram")
plt.xlabel("Numbers")
plt.ylabel("Frequency")
plt.show()



# labels = 'Frogs', 'Hogs', 'Dogs', 'Logs'
# sizes = [15, 30, 45, 10]
# explode = (0, 0.1, 0, 0)
# plt.pie(sizes, explode=explode, labels=labels)
# plt.show()


t = arange(-1.0, 1.0, 0.01)
s = sin(2.5 * pi * t)
plot(t,s)
title("sine wave")
xlabel("Time")
ylabel("Amplitude")
show()

s = cos(2.5 * pi * t)
plot(t,s)
title("cos wave")
xlabel("Time")
ylabel("Amplitude")
show()
# SToring all sheets with results of all questions in excel file 173050061_solution.xls
sh = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(sh)
wb.save(filename = dest_filename)

print ("Success..... " + "\n" + "check Output files....")